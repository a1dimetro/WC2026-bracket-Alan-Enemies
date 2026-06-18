#!/usr/bin/env python3
"""
Automated leaderboard updater for the World Cup 2026 bracket pool.

This script scans a directory of bracket spreadsheets (Excel files), reads each
participant's predictions from the `Fixture` sheet and compares them to the
actual match results provided in a CSV file.  It awards points using the
standard World Cup pool scoring system (1 point for correct outcome,
another point for correct goal difference, and a third point for the exact
scoreline) and outputs a JSON file containing the total points per bracket.

The script is designed to run on GitHub Actions or locally.  It assumes
that the bracket files are stored in a directory called `brackets/` and
the results file is named `results.csv` in the repository root.  The
output JSON (`leaderboard.json`) will be written to the repository root.

Usage (from repo root):
    python scripts/update_leaderboard.py

Environment variables:
    BRACKETS_DIR   Directory containing bracket spreadsheets (default: brackets)
    RESULTS_FILE   Path to CSV file with match results (default: results.csv)
    OUTPUT_FILE    Path to output JSON file (default: leaderboard.json)

The results CSV must contain the following columns:
    date        Date of the match (YYYY-MM-DD) – not used in scoring but kept for reference
    group       Group letter or round descriptor
    home_team   Full name of the home team (matching FIFA fixtures page)
    away_team   Full name of the away team
    home_score  Integer goals scored by the home team
    away_score  Integer goals scored by the away team

You must also provide a mapping from team full names to the three-letter
codes used in the bracket spreadsheets.  This mapping can be customised
below (TEAM_CODE_MAP).  If a team name appears in the results but is not
present in this map, the script will raise an exception.
"""

import json
import os
import sys
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# Mapping from full team names (used in results.csv) to the three-letter codes
# used in the bracket spreadsheets.  Extend this map as necessary when new
# teams appear in the tournament.
# Mapping from full team names (used in results.csv/results.xlsx) to the three-letter codes
# used in the bracket spreadsheets.  Extend this map as necessary when new
# teams appear in the tournament.
TEAM_CODE_MAP = {
    'Mexico': 'Mex',
    'South Africa': 'Sou',
    'South Korea': 'Cor',  # South Korea is sometimes abbreviated as "Cor" for Corea
    # Some sources list the country as "Czech Republic" while others use "Czechia".
    'Czech Republic': 'Cze',
    'Czechia': 'Cze',
    'Canada': 'Can',
    'Bosnia and Herzegovina': 'Bos',
    'Qatar': 'Qat',
    'Switzerland': 'Swi',
    'Brazil': 'Bra',
    'Morocco': 'Mor',
    'Haiti': 'Hai',
    'Scotland': 'Sco',
    'United States': 'Uni',  # USA appears as "Uni" in spreadsheets
    'Paraguay': 'Par',
    'Australia': 'Aus',
    'Turkey': 'Tur',
    # The FIFA fixtures page often uses the native spelling "Turkiye".
    'Turkiye': 'Tur',
    'Germany': 'Ger',
    'Curaçao': 'Cur',
    'Ivory Coast': 'Ivo',  # Côte d'Ivoire shortened to Ivo
    "Côte d'Ivoire": 'Ivo',  # accented version
    'Croatia': 'Cro',
    'Ecuador': 'Ecu',
    'England': 'Eng',
    'Korea Republic': 'Cor',
    'Japan': 'Jap',
    'Netherlands': 'Net',
    'Sweden': 'Swe',
    'Tunisia': 'Tun',
    # Additional teams for other groups (extend as needed)
    'Cape Verde': 'Cap',
    'Spain': 'Spa',
    'Belgium': 'Bel',
    'Egypt': 'Egy',
    'Iran': 'Ira',
    'New Zealand': 'New',
    'Saudi Arabia': 'Sau',
    'Uruguay': 'Uru',
    'France': 'Fra',
    'Senegal': 'Sen',
    'Iraq': 'Irq',
    'Norway': 'Nor',
    'Argentina': 'Arg',
    'Algeria': 'Alg',
    # Austria and Australia both use the code "Aus" in the bracket spreadsheets.
    'Austria': 'Aus',
    'Jordan': 'Jor',
    'Portugal': 'Por',
    'DR Congo': 'Drc',
    'Ghana': 'Gha',
    'Panama': 'Pan',
    # Some results contain a typo "Uzbakistan"; map both to Uzb.
    'Uzbekistan': 'Uzb',
    'Uzbakistan': 'Uzb',
    'Colombia': 'Col',
    # Additional synonyms or alternate country spellings can be added here as needed
}


def load_results(results_file: str) -> list:
    """Load match results from a CSV or Excel file.

    The results file may be either a CSV (comma‑separated) or an Excel workbook.
    The expected columns are:
        date, group, home_team, away_team, home_score, away_score

    Returns a list of tuples: (home_code, away_code, home_score, away_score).
    """
    # Determine file type based on extension
    _, ext = os.path.splitext(results_file)
    if ext.lower() in {'.xlsx', '.xlsm', '.xls'}:
        results_df = pd.read_excel(results_file, engine='openpyxl')
    else:
        results_df = pd.read_csv(results_file)
    results = []
    for _, row in results_df.iterrows():
        home_team = row['home_team']
        away_team = row['away_team']
        if home_team not in TEAM_CODE_MAP or away_team not in TEAM_CODE_MAP:
            raise ValueError(
                f"Team mapping missing for '{home_team}' or '{away_team}'. Please update TEAM_CODE_MAP."
            )
        home_code = TEAM_CODE_MAP[home_team]
        away_code = TEAM_CODE_MAP[away_team]
        results.append(
            (home_code, away_code, int(row['home_score']), int(row['away_score']))
        )
    return results


def parse_bracket_predictions(file_path: str) -> dict:
    """Parse predicted match scores from a bracket spreadsheet.

    This function reads the `WORLDCUP` sheet and extracts predictions from
    columns AA (home team), AC (home goals), AD (away goals), and AF
    (away team).  These columns correspond to the two teams and their
    predicted scores.  Some spreadsheets may leave blank columns (AB,
    AE) between these, so we reference the specific letters.

    Returns a dictionary mapping (home_code, away_code) to a tuple of
    (pred_home_goals, pred_away_goals).  Only the earliest occurrence of
    each fixture is kept (subsequent rounds of the same pairing are
    ignored).  Predictions with missing scores or unmapped team names
    are skipped.
    """
    predictions: dict = {}
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return predictions
    if 'WORLDCUP' not in wb.sheetnames:
        print(f"Sheet 'WORLDCUP' not found in {file_path}.")
        return predictions
    ws = wb['WORLDCUP']
    # Determine column indices for team names and scores
    idx_team1 = column_index_from_string('AA')  # home team name
    idx_home_score = column_index_from_string('AC')  # home goals
    idx_away_score = column_index_from_string('AD')  # away goals
    idx_team2 = column_index_from_string('AF')  # away team name
    for row in ws.iter_rows(min_row=1):
        # openpyxl rows are tuples of cells; convert to values
        team1 = row[idx_team1 - 1].value
        home_goals = row[idx_home_score - 1].value
        away_goals = row[idx_away_score - 1].value
        team2 = row[idx_team2 - 1].value
        # Skip header rows or rows without both team names and numeric scores
        if not team1 or not team2:
            continue
        # Skip rows where scores are not numeric (e.g., 'Goals' headers)
        if not isinstance(home_goals, (int, float)) or not isinstance(away_goals, (int, float)):
            continue
        # Map team names to codes
        team1 = str(team1).strip()
        team2 = str(team2).strip()
        if team1 not in TEAM_CODE_MAP or team2 not in TEAM_CODE_MAP:
            # Unknown team names are ignored
            continue
        home_code = TEAM_CODE_MAP[team1]
        away_code = TEAM_CODE_MAP[team2]
        pred_home_goals = int(home_goals)
        pred_away_goals = int(away_goals)
        key = (home_code, away_code)
        if key not in predictions:
            predictions[key] = (pred_home_goals, pred_away_goals)
    return predictions


def compute_points(pred: tuple, actual: tuple) -> int:
    """Compute points for a single fixture prediction.

    pred and actual are tuples (home_goals, away_goals).
    """
    pred_home, pred_away = pred
    act_home, act_away = actual
    # Determine outcomes
    def outcome(goals_home: int, goals_away: int) -> str:
        return 'home' if goals_home > goals_away else 'away' if goals_home < goals_away else 'draw'
    pred_outcome = outcome(pred_home, pred_away)
    act_outcome = outcome(act_home, act_away)
    points = 0
    # Correct outcome
    if pred_outcome == act_outcome:
        points = 1
        # Correct goal difference
        if (pred_home - pred_away) == (act_home - act_away):
            points = 2
            # Exact score
            if pred_home == act_home and pred_away == act_away:
                points = 3
    return points


def main():
    brackets_dir = os.environ.get('BRACKETS_DIR', 'brackets')
    results_file = os.environ.get('RESULTS_FILE', 'results.csv')
    output_file = os.environ.get('OUTPUT_FILE', 'leaderboard.json')
    if not os.path.exists(results_file):
        print(f"Results file '{results_file}' not found. Please provide a CSV of match results.")
        sys.exit(1)
    if not os.path.isdir(brackets_dir):
        print(f"Brackets directory '{brackets_dir}' not found.")
        sys.exit(1)
    # Load actual results
    results = load_results(results_file)
    # Build a map of actual results for quick lookup
    results_map = {(home, away): (home_score, away_score) for home, away, home_score, away_score in results}
    # To account for reversed fixtures (if predictions list team order differently),
    # also store reversed mapping
    reversed_map = {(away, home): (away_score, home_score) for home, away, home_score, away_score in results}
    # Initialize scoreboard
    leaderboard = []
    for filename in sorted(os.listdir(brackets_dir)):
        if filename.lower().endswith('.xlsx'):
            file_path = os.path.join(brackets_dir, filename)
            bracket_name = os.path.splitext(filename)[0]
            preds = parse_bracket_predictions(file_path)
            total_points = 0
            for matchup, pred_score in preds.items():
                # Check if actual result exists (direct or reversed)
                act_score = results_map.get(matchup)
                if act_score is None:
                    act_score = reversed_map.get(matchup)
                    if act_score is None:
                        # Match not yet played or not in results file
                        continue
                # Compute points
                total_points += compute_points(pred_score, act_score)
            leaderboard.append({
                'bracket': bracket_name,
                'total_points': total_points,
            })
    # Sort leaderboard by total points descending
    leaderboard.sort(key=lambda x: (-x['total_points'], x['bracket']))
    # Write output JSON
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(leaderboard, f, indent=2)
    print(f"Leaderboard updated with {len(leaderboard)} brackets. Results written to {output_file}.")


if __name__ == '__main__':
    main()